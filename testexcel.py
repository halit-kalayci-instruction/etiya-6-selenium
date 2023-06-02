import openpyxl
def readDataFromExcel():
        excelFile = openpyxl.load_workbook("data/userFailData.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row # sayfadaki max. satırı alır.
        data = []
        for i in range(1,rows):
            username = sheet.cell(i, 1).value
            password = sheet.cell(i, 2).value
            data.append((username,password))
        return data

readDataFromExcel()